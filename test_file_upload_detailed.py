#!/usr/bin/env python3
"""
Script detallado para probar la carga de archivos y verificar el estado
"""
import requests
import os
import sys
import json

def test_file_upload_detailed():
    """Prueba detallada de carga de archivos"""
    
    base_url = "https://calculadora-fabricabilidad.onrender.com"
    session = requests.Session()
    
    print("🔍 Iniciando prueba detallada de carga de archivos...")
    print(f"🌐 URL: {base_url}")
    print("-" * 50)
    
    # Paso 1: Login
    print("1️⃣ Haciendo login...")
    login_data = {
        'username': 'admin',
        'password': 'admin123'
    }
    
    try:
        response = session.post(f"{base_url}/login", data=login_data, allow_redirects=False)
        if response.status_code == 302:
            print("   ✅ Login exitoso")
            # Seguir redirección
            redirect_url = response.headers.get('Location')
            if redirect_url:
                response = session.get(f"{base_url}{redirect_url}")
        else:
            print(f"   ❌ Error en login: {response.status_code}")
            return False
    except Exception as e:
        print(f"   ❌ Error en login: {e}")
        return False
    
    # Paso 2: Verificar estado inicial
    print("\n2️⃣ Verificando estado inicial...")
    try:
        response = session.get(base_url)
        if response.status_code == 200:
            print("   ✅ Página accesible")
            # Buscar badges de estado en el HTML
            html_content = response.text
            if 'historico_costos_loaded' in html_content:
                print("   ✅ Variable historico_costos_loaded encontrada en HTML")
            else:
                print("   ⚠️  Variable historico_costos_loaded NO encontrada en HTML")
        else:
            print(f"   ❌ Error accediendo a página: {response.status_code}")
            return False
    except Exception as e:
        print(f"   ❌ Error verificando estado: {e}")
        return False
    
    # Paso 3: Crear archivo de prueba más realista
    print("\n3️⃣ Creando archivo de prueba realista...")
    
    # Crear un archivo Excel real usando openpyxl
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico Costos"
        
        # Agregar encabezados
        ws['A1'] = "Artículo"
        ws['B1'] = "2023"
        ws['C1'] = "2024"
        
        # Agregar datos
        ws['A2'] = "A001"
        ws['B2'] = "10.00"
        ws['C2'] = "10.50"
        
        ws['A3'] = "A002"
        ws['B3'] = "15.00"
        ws['C3'] = "15.75"
        
        # Guardar archivo
        test_filename = "test_historico_real.xlsx"
        wb.save(test_filename)
        print(f"   ✅ Archivo creado: {test_filename}")
        
    except ImportError:
        print("   ⚠️  openpyxl no disponible, creando archivo simple...")
        # Crear archivo simple como fallback
        with open('test_historico_real.xlsx', 'w') as f:
            f.write("Artículo\t2023\t2024\nA001\t10.00\t10.50\nA002\t15.00\t15.75")
    
    # Paso 4: Cargar archivo y verificar respuesta
    print("\n4️⃣ Cargando archivo de histórico de costos...")
    
    try:
        with open(test_filename, 'rb') as f:
            files = {
                'historico_costos_file': (test_filename, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            response = session.post(base_url, files=files, allow_redirects=False)
            
            print(f"   Status: {response.status_code}")
            print(f"   Headers: {dict(response.headers)}")
            
            if response.status_code == 302:
                print("   ✅ Archivo procesado (redirección)")
                
                # Seguir la redirección y verificar el estado
                redirect_url = response.headers.get('Location')
                if redirect_url:
                    print(f"   🔄 Siguiendo redirección a: {redirect_url}")
                    response = session.get(f"{base_url}{redirect_url}")
                    
                    if response.status_code == 200:
                        html_content = response.text
                        
                        # Buscar indicadores de éxito
                        if 'Archivo "test_historico_real.xlsx" procesado correctamente' in html_content:
                            print("   ✅ Mensaje de éxito encontrado en HTML")
                        else:
                            print("   ⚠️  Mensaje de éxito NO encontrado en HTML")
                        
                        # Buscar badge de estado
                        if 'badge-success' in html_content and 'Cargado' in html_content:
                            print("   ✅ Badge de 'Cargado' encontrado")
                        else:
                            print("   ⚠️  Badge de 'Cargado' NO encontrado")
                        
                        # Buscar variable de estado
                        if 'historico_costos_loaded' in html_content:
                            print("   ✅ Variable historico_costos_loaded presente")
                        else:
                            print("   ⚠️  Variable historico_costos_loaded NO presente")
                        
                        # Mostrar fragmento del HTML para debug
                        print("\n   📄 Fragmento del HTML (últimas 500 caracteres):")
                        print("   " + "-" * 50)
                        print(html_content[-500:])
                        print("   " + "-" * 50)
                        
                    else:
                        print(f"   ❌ Error en redirección: {response.status_code}")
            else:
                print(f"   ❌ Error procesando archivo: {response.status_code}")
                print(f"   Respuesta: {response.text[:300]}...")
                
    except Exception as e:
        print(f"   ❌ Error cargando archivo: {e}")
    
    # Limpiar archivo de prueba
    print("\n5️⃣ Limpiando archivo de prueba...")
    if os.path.exists(test_filename):
        os.remove(test_filename)
        print(f"   🗑️  Eliminado: {test_filename}")
    
    print("\n✅ Prueba detallada completada")
    return True

if __name__ == "__main__":
    test_file_upload_detailed() 