import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import pandas as pd # Importar pandas para manejar DataFrames

# Importamos las clases correctas y existentes de data_models
from data_models import IndividualCalculationResult, PurchaseSuggestion, ComponentDetail, LotCalculationResult, DemandProjectionResult, ComponentDemandDetail, PurchaseSuggestionDemand, EqualizationComponentSummary, EqualizationResult, ModelFullCostResult # Asegúrate de que todas las dataclasses estén importadas

class ReportGenerator:
    def __init__(self, data):
        self.data = data
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def _apply_header_styles(self, cell_range):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in self.sheet[cell_range]:
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    def _generate_individual_report(self):
        self.sheet.title = "Informe Individual"
        
        # --- Resumen ---
        self.sheet['A1'] = "Resumen del Cálculo para Modelo:"
        self.sheet['B1'] = self.data['nombre_modelo']
        self.sheet['A1'].font = Font(bold=True)
        
        headers = ["Cantidad Deseada", "Cantidad Fabricable", "Costo Total Sugerencias ($)"]
        values = [self.data['desired_qty'], self.data['cantidad_fabricable'], self.data['costo_total_sugerencias']]
        
        # Escribir encabezados y valores en filas separadas para mejor presentación
        self.sheet.append([]) # Fila vacía para espacio
        self.sheet.append(headers)
        self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(headers))}{self.sheet.max_row}')
        self.sheet.append(values)
        
        # Formato de moneda para el costo total
        self.sheet[f'C{self.sheet.max_row}'].number_format = '"$"#,##0.00'

        # --- Sugerencias de Compra ---
        if self.data['sugerencias_compra']:
            self.sheet.append([]) # Fila vacía
            self.sheet.append([]) # Fila vacía
            self.sheet.cell(row=self.sheet.max_row, column=1, value="Sugerencias de Compra de Componentes").font = Font(bold=True)
            purchase_headers = ["Artículo", "Cantidad Necesaria", "Costo Unitario ($)", "Costo Total ($)"]
            self.sheet.append(purchase_headers)
            self._apply_header_styles(f'A{self.sheet.max_row}:D{self.sheet.max_row}')
            
            for s in self.data['sugerencias_compra']:
                row_data = [s['articulo'], s['cantidad_necesaria'], s['costo_unitario'], s['costo_total']]
                self.sheet.append(row_data)
                # Formato de moneda para columnas de costo
                self.sheet[f'C{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                self.sheet[f'D{self.sheet.max_row}'].number_format = '"$"#,##0.00'
        else:
            self.sheet.append([]) # Fila vacía
            self.sheet.append([]) # Fila vacía
            self.sheet.cell(row=self.sheet.max_row, column=1, value="🎉 ¡Excelente! No se necesitan compras adicionales para fabricar la cantidad deseada.").font = Font(italic=True, color="008000")


        # --- Detalle de Componentes ---
        if self.data['detalle_componentes']:
            self.sheet.append([]) # Fila vacía
            self.sheet.append([]) # Fila vacía
            self.sheet.cell(row=self.sheet.max_row, column=1, value="Detalle de Componentes").font = Font(bold=True)
            detail_headers = ["Artículo", "Descripción", "Cant. Requerida (Total)", "Stock Disponible", "Cant. Faltante (p/ deseada)", "Costo Unitario ($)", "Costo Total ($)"]
            self.sheet.append(detail_headers)
            detail_start_row = self.sheet.max_row # Fila donde inician los encabezados del detalle
            self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(detail_headers))}{self.sheet.max_row}')
            
            for d in self.data['detalle_componentes']:
                row_data = [d['articulo'], d['articulo_descripcion'], d['cantidad_requerida_total'],
                            d['cantidad_disponible_stock'], d['cantidad_faltante'], d['costo_unitario'], d['costo_total']]
                self.sheet.append(row_data)
                # Formato de moneda para columnas de costo
                self.sheet[f'F{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                self.sheet[f'G{self.sheet.max_row}'].number_format = '"$"#,##0.00'
            
            detail_end_row = self.sheet.max_row # Última fila de datos del detalle

            # --- Añadir Gráfico de Cantidad Faltante ---
            # Solo si hay componentes faltantes
            faltantes_data = [d for d in self.data['detalle_componentes'] if d['cantidad_faltante'] > 0]
            if faltantes_data:
                chart = BarChart()
                chart.type = "col" # Gráfico de columnas
                chart.style = 10 # Estilo de gráfico
                chart.title = "Cantidad Faltante por Componente"
                chart.y_axis.title = "Cantidad Faltante"
                chart.x_axis.title = "Artículo"

                # Definir los datos para el gráfico
                # Columna 'Cant. Faltante (p/ deseada)' es la 5ta columna (índice 4) en el detalle_headers
                data_ref = Reference(self.sheet, min_col=5, min_row=detail_start_row + 1, max_row=detail_end_row)
                # Columna 'Artículo' es la 1ra columna (índice 0) en el detalle_headers
                categories_ref = Reference(self.sheet, min_col=1, min_row=detail_start_row + 1, max_row=detail_end_row)

                chart.add_data(data_ref, titles_from_data=False)
                chart.set_categories(categories_ref)

                # Ajustar el tamaño del gráfico
                chart.width = 15 # Ancho en unidades de Excel
                chart.height = 10 # Alto en unidades de Excel

                # Posicionar el gráfico (ej. en la columna H, después de las tablas)
                self.sheet.add_chart(chart, f"H{detail_start_row}")

        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width

    def _generate_lot_report(self):
        self.sheet.title = "Informe de Lote"

        # Resumen de resultados por modelo en el lote
        current_row = 1
        for model_code, model_data in self.data['results'].items():
            self.sheet.cell(row=current_row, column=1, value=f"Modelo: {model_data['nombre_modelo']}").font = Font(bold=True, size=14)
            current_row += 1
            self.sheet.cell(row=current_row, column=1, value="Cantidad Deseada:").font = Font(bold=True)
            self.sheet.cell(row=current_row, column=2, value=model_data['desired_qty'])
            current_row += 1
            self.sheet.cell(row=current_row, column=1, value="Cantidad Fabricable:").font = Font(bold=True)
            self.sheet.cell(row=current_row, column=2, value=model_data['cantidad_fabricable'])
            current_row += 1
            self.sheet.cell(row=current_row, column=1, value="Costo Total Sugerencias:").font = Font(bold=True)
            self.sheet.cell(row=current_row, column=2, value=model_data['costo_total_sugerencias']).number_format = '"$"#,##0.00'
            current_row += 2 # Espacio entre modelos

            # Detalle de componentes para cada modelo
            if model_data['detalle_componentes']:
                self.sheet.cell(row=current_row, column=1, value="Detalle de Componentes").font = Font(bold=True)
                current_row += 1
                detail_headers = ["Artículo", "Descripción", "Cant. Requerida (Total)", "Stock Disponible", "Cant. Faltante (p/ deseada)", "Costo Unitario ($)", "Costo Total ($)"]
                self.sheet.append(detail_headers)
                self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(detail_headers))}{self.sheet.max_row}')
                current_row += 1

                for d in model_data['detalle_componentes']:
                    row_data = [d['articulo'], d['articulo_descripcion'], d['cantidad_requerida_total'],
                                d['cantidad_disponible_stock'], d['cantidad_faltante'], d['costo_unitario'], d['costo_total']]
                    self.sheet.append(row_data)
                    self.sheet[f'F{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                    self.sheet[f'G{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                    current_row += 1
                current_row += 1 # Espacio después de la tabla de detalle
            
            # Sugerencias de compra para cada modelo (si existen)
            if model_data['sugerencias_compra']:
                self.sheet.cell(row=current_row, column=1, value="Sugerencias de Compra para este Modelo").font = Font(bold=True)
                current_row += 1
                purchase_headers = ["Artículo", "Cantidad Necesaria", "Costo Unitario ($)", "Costo Total ($)"]
                self.sheet.append(purchase_headers)
                self._apply_header_styles(f'A{self.sheet.max_row}:D{self.sheet.max_row}')
                current_row += 1

                for s in model_data['sugerencias_compra']:
                    row_data = [s['articulo'], s['cantidad_necesaria'], s['costo_unitario'], s['costo_total']]
                    self.sheet.append(row_data)
                    self.sheet[f'C{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                    self.sheet[f'D{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                    current_row += 1
                current_row += 2 # Espacio después de las sugerencias

        # Sugerencias de compra agregadas para todo el lote
        if self.data['suggestions']:
            self.sheet.append([]) # Fila vacía
            self.sheet.append([]) # Fila vacía
            self.sheet.cell(row=self.sheet.max_row, column=1, value="Sugerencias de Compra Agregadas para el Lote").font = Font(bold=True, size=14)
            current_row = self.sheet.max_row + 1
            
            purchase_headers = ["Artículo", "Cantidad Necesaria Agregada", "Costo Unitario ($)", "Costo Total Estimado ($)"]
            self.sheet.append(purchase_headers)
            self._apply_header_styles(f'A{self.sheet.max_row}:D{self.sheet.max_row}')
            current_row = self.sheet.max_row + 1

            for item_code, item in self.data['suggestions'].items():
                row_data = [item['articulo'], item['cantidad_necesaria'], item['costo_unitario'], item['costo_total']]
                # Asegurarse de que los valores sean numéricos para el formato
                row_data[1] = float(row_data[1])
                row_data[2] = float(row_data[2])
                row_data[3] = float(row_data[3])

                self.sheet.append(row_data)
                self.sheet[f'C{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                self.sheet[f'D{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                current_row += 1
        
        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width

    def _generate_demand_report(self):
        self.sheet.title = "Informe de Demanda"

        # Usar notación de punto para DemandProjectionResult
        self.sheet['A1'] = f"Sugerencias de Compra (Proyección a {self.data.projection_period} Meses)"
        self.sheet['A1'].font = Font(bold=True, size=14)
        
        current_row = 3 # Empezar después del título

        # Si self.data es un DemandProjectionResult, convertir a dict para el resto del acceso
        data_dict = self.data.to_dict() if hasattr(self.data, 'to_dict') else self.data

        if data_dict.get('mensaje'):
            self.sheet.cell(row=current_row, column=1, value=data_dict['mensaje']).font = Font(italic=True, color="0000FF")
            current_row += 2
        elif not data_dict.get('sugerencias_agrupadas'):
            self.sheet.cell(row=current_row, column=1, value="No se generaron sugerencias de compra. Esto puede deberse a que el stock actual ya cubre la demanda proyectada.").font = Font(italic=True, color="0000FF")
            current_row += 2
        else:
            for modelo, data in data_dict['sugerencias_agrupadas'].items():
                if data['componentes_necesarios']:
                    self.sheet.cell(row=current_row, column=1, value=f"Modelo: {modelo}").font = Font(bold=True, size=12)
                    current_row += 1
                    # La demanda proyectada del modelo ya viene formateada como string en app.py
                    self.sheet.cell(row=current_row, column=1, value=f"Demanda Proyectada para este Modelo: {data['demanda_proyectada_modelo']} unidades").font = Font(italic=True)
                    current_row += 2

                    headers = ["Artículo (Componente)", "Stock Actual", "Necesidad Proyectada", "Cantidad a Comprar", "Costo de Compra ($)"]
                    self.sheet.append(headers)
                    self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(headers))}{self.sheet.max_row}')
                    current_row += 1

                    for s in data['componentes_necesarios']:
                        # Los valores ya vienen formateados como string desde app.py
                        row_data = [s['articulo'], s['stock_actual'], s['demanda_proyectada_componente'], s['cantidad_a_comprar'], s['costo_total_compra']]
                        self.sheet.append(row_data)
                        # No aplicar formato numérico aquí, ya que los valores son strings formateados
                        current_row += 1
                    current_row += 2 # Espacio entre modelos

        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width

    def _generate_equalization_report(self):
        self.sheet.title = "Informe de Equilibrado de Stock"

        self.sheet['A1'] = "Resultados de Equilibrado de Stock"
        self.sheet['A1'].font = Font(bold=True, size=14)
        
        current_row = 3 # Empezar después del título

        # Si self.data es un EqualizationResult, convertir a dict para el resto del acceso
        data_dict = self.data.to_dict() if hasattr(self.data, 'to_dict') else self.data

        # Mostrar mensaje solo si NO hay componentes
        if not data_dict.get('component_summaries') or len(data_dict.get('component_summaries', [])) == 0:
            msg = data_dict.get('message', 'No se generaron resúmenes de componentes para el equilibrado de stock.')
            self.sheet.cell(row=current_row, column=1, value=msg).font = Font(italic=True, color="0000FF")
            current_row += 2
        else:
            self.sheet.cell(row=current_row, column=1, value=f"Período de Proyección: {data_dict['projection_period_months']} Meses").font = Font(bold=True)
            current_row += 1
            self.sheet.cell(row=current_row, column=1, value=f"Costo Total de Compra Después de Equilibrado: {data_dict['total_cost_after_equalization']}").font = Font(bold=True)
            current_row += 2

            headers = [
                "Artículo", "Descripción", "Stock Actual Global", "Demanda Proyectada Global", "Excedente Global", "Déficit Global",
                "Proveedor Sugerido", "Código Proveedor", "Costo Unitario ($)", "Cantidad a Comprar Final", "Costo Total Compra Final ($)"
            ]
            self.sheet.append(headers)
            self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(headers))}{self.sheet.max_row}')
            current_row += 1

            for s in data_dict['component_summaries']:
                row_data = [
                    s.get('articulo', ''),
                    s.get('articulo_descripcion', ''),
                    s.get('stock_disponible', ''), s.get('demanda_total', ''),
                    s.get('excedente_global', ''),
                    s.get('cantidad_faltante_original', ''),
                    s.get('razon_social_proveedor_final', ''),
                    s.get('codigo_proveedor_final', ''),
                    s.get('costo_unitario_proveedor_final', ''),
                    s.get('cantidad_a_comprar_final', ''), s.get('costo_total_compra_final', '')
                ]
                self.sheet.append(row_data)
                current_row += 1
            current_row += 2 # Espacio

        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width

    def _generate_suppliers_report(self):
        """
        Genera un informe Excel para los datos de proveedores.
        Se espera que self.data sea un DataFrame de pandas.
        """
        self.sheet.title = "Informe de Proveedores"

        if self.data.empty:
            self.sheet['A1'] = "No hay datos de proveedores para generar el informe."
            self.sheet['A1'].font = Font(italic=True, color="FF0000")
            return

        # Escribir encabezados
        headers = self.data.columns.tolist()
        self.sheet.append(headers)
        self._apply_header_styles(f'A1:{get_column_letter(len(headers))}1')

        # Escribir datos
        for r_idx, row in self.data.iterrows():
            row_data = row.tolist()
            self.sheet.append(row_data)
            # Aplicar formato de moneda a la columna 'precio'
            try:
                precio_col_idx = headers.index('precio') + 1 # +1 porque openpyxl es 1-based
                self.sheet.cell(row=self.sheet.max_row, column=precio_col_idx).number_format = '"$"#,##0.00'
            except ValueError:
                pass # La columna 'precio' no se encontró, no aplicar formato

        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width

    def _generate_model_full_cost_report(self):
        """
        Genera un informe Excel para el costo total de fabricación de un modelo.
        Se espera que self.data sea un diccionario que representa un ModelFullCostResult.
        """
        self.sheet.title = "Costo Total Modelo"

        if self.data.get('mensaje'):
            self.sheet['A1'] = self.data['mensaje']
            self.sheet['A1'].font = Font(italic=True, color="FF0000")
            return

        self.sheet['A1'] = "Costo Total de Fabricación para Modelo:"
        self.sheet['B1'] = self.data['nombre_modelo']
        self.sheet['A1'].font = Font(bold=True)

        self.sheet['A2'] = "Cantidad del Modelo:"
        self.sheet['B2'] = self.data['cantidad_modelo']
        self.sheet['A2'].font = Font(bold=True)

        self.sheet['A3'] = "Costo Total de Fabricación:"
        self.sheet['B3'] = self.data['costo_total_fabricacion']
        self.sheet['B3'].number_format = '"$"#,##0.00'
        self.sheet['A3'].font = Font(bold=True)
        
        current_row = 5 # Empezar después del resumen

        if self.data['detalle_componentes']:
            self.sheet.cell(row=current_row, column=1, value="Detalle de Componentes por Costo").font = Font(bold=True)
            current_row += 1
            detail_headers = ["Artículo", "Descripción", "Cantidad Requerida (Total)", "Costo Unitario ($)", "Costo Total ($)"]
            self.sheet.append(detail_headers)
            self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(detail_headers))}{self.sheet.max_row}')
            current_row += 1

            for d in self.data['detalle_componentes']:
                row_data = [d['articulo'], d['articulo_descripcion'], d['cantidad_requerida_total'],
                            d['costo_unitario'], d['costo_total']]
                self.sheet.append(row_data)
                self.sheet[f'D{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                self.sheet[f'E{self.sheet.max_row}'].number_format = '"$"#,##0.00'
                current_row += 1
        
        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width

    def _generate_costs_report(self):
        """
        Genera un informe Excel para el histórico de costos.
        Se espera que self.data sea un DataFrame de pandas con los datos de costos históricos.
        """
        self.sheet.title = "Histórico de Costos"

        if self.data.empty:
            self.sheet['A1'] = "No hay datos de costos históricos disponibles."
            self.sheet['A1'].font = Font(italic=True, color="FF0000")
            return

        # Título del reporte
        self.sheet['A1'] = "Reporte de Costos Históricos"
        self.sheet['A1'].font = Font(bold=True, size=14)
        self.sheet.merge_cells('A1:D1')

        # Información del reporte
        self.sheet['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        self.sheet['A3'].font = Font(italic=True)

        # Encabezados de la tabla
        headers = list(self.data.columns)
        self.sheet.append([])  # Fila vacía
        self.sheet.append(headers)
        self._apply_header_styles(f'A{self.sheet.max_row}:{get_column_letter(len(headers))}{self.sheet.max_row}')

        # Datos
        for index, row in self.data.iterrows():
            row_data = []
            for col in self.data.columns:
                value = row[col]
                if pd.isna(value):
                    row_data.append("")
                else:
                    row_data.append(value)
            self.sheet.append(row_data)

        # Ajustar ancho de columnas automáticamente
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[column].width = adjusted_width


    def generate_report(self, report_type):
        if report_type == 'individual':
            self._generate_individual_report()
        elif report_type == 'lot':
            self._generate_lot_report()
        elif report_type == 'demand':
            self._generate_demand_report()
        elif report_type == 'equalization': # Añadir el nuevo tipo de reporte de equilibrado
            self._generate_equalization_report()
        elif report_type == 'suppliers': # Nuevo tipo de reporte para proveedores
            self._generate_suppliers_report()
        elif report_type == 'model_full_cost': # NUEVO TIPO DE REPORTE
            self._generate_model_full_cost_report()
        elif report_type == 'costs': # NUEVO TIPO DE REPORTE PARA COSTOS HISTÓRICOS
            self._generate_costs_report()
        else:
            raise ValueError(f"Tipo de reporte '{report_type}' no soportado.")

        # Guardar el archivo temporalmente para enviarlo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Reporte_{report_type}_{timestamp}.xlsx"
        
        # Guardar en una carpeta 'reports'
        reports_folder = 'reports'
        if not os.path.exists(reports_folder):
            os.makedirs(reports_folder)
        
        file_path = os.path.join(reports_folder, file_name)
        self.workbook.save(file_path)
        return file_path
